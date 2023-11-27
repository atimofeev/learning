import openpyxl as o

RED = "\033[91m"
GREEN = "\033[92m"
RESET = "\033[0m"


INPUT_FILE = "spreadsheet.xlsx"
WORKBOOK = o.load_workbook(INPUT_FILE)

worksheets = [WORKBOOK[sheetname] for sheetname in WORKBOOK.sheetnames]

def show_menu():
    """Function to show menu"""
    print("This is some spreadsheet tool")
    print(f"{GREEN}1.{RESET} Print all sheets")
    print(f"{GREEN}2.{RESET} Find low quantity of sales")
    print(f"{GREEN}3.{RESET} Print margin from Sales Data")
    print(f"{GREEN}4.{RESET} Write margin to Sales Data")
    print(f"{GREEN}5.{RESET} Remove margin from Sales Data")
    print(f"{GREEN}6.{RESET} Check Sales Data")

def print_worksheet(worksheet):
    """Function to print content of a single worksheet"""
    # Determine the maximum width for each column
    col_widths = [max(len(str(cell.value)) for cell in col) for col in worksheet.iter_cols()]

    for row in worksheet.iter_rows():
        for cell, width in zip(row, col_widths):
            # Print each cell value, left-aligned and padded to the column width
            print(str(cell.value).ljust(width), end="\t")
        print()  # Print a newline at the end of each row

def print_sheets(sheet_index=None):
    """Function to print specified sheet or all sheets if no index is provided"""
    if sheet_index is not None:
        print_worksheet(worksheets[sheet_index])
    else:
        for worksheet in worksheets:
            print_worksheet(worksheet)

def print_margin():
    """Function calculating margin from product sheet"""
    product_margin = {}
    worksheet = worksheets[3]

    for row in range(2, worksheet.max_row +1):
        product_name = worksheet.cell(row, 2).value
        quantity_sold = int(worksheet.cell(row, 3).value)
        price_per_unit = int(worksheet.cell(row, 4).value)
        cost_per_unit = int(worksheet.cell(row, 5).value)
        margin = (price_per_unit - cost_per_unit) * quantity_sold

        if product_name in product_margin:
            product_margin[product_name] += margin
        else:
            product_margin[product_name] = margin

    for product, margin in product_margin.items():
        print(f"Margin for {product}: {GREEN}{margin}{RESET}")

def find_low_quantity():
    """Function to find low quantity of sales"""
    product_sales = {}
    worksheet = worksheets[3]
    print(f"List of products with sales < {RED}10{RESET}:")

    for row in range(2, worksheet.max_row +1):
        product_name = worksheet.cell(row, 2).value
        quantity_sold = int(worksheet.cell(row, 3).value)

        if quantity_sold < 10:

            if product_name in product_sales:
                product_sales[product_name] += quantity_sold
            else:
                product_sales[product_name] = quantity_sold

    for product, sales in product_sales.items():
        print(f"{product}: {RED}{sales}{RESET}")

def write_margin():
    """Function to write margin value to each row"""
    worksheet = worksheets[3]
    worksheet.cell(1, worksheet.max_column + 1).value = "Margin"

    for row in range(2, worksheet.max_row +1):
        price_per_unit = int(worksheet.cell(row, 4).value)
        cost_per_unit = int(worksheet.cell(row, 5).value)
        worksheet.cell(row, worksheet.max_column).value = price_per_unit - cost_per_unit

    WORKBOOK.save(INPUT_FILE)
    print(f"{GREEN}File updated!{RESET}")

def delete_margin():
    """Function to remove margin from each row"""
    worksheet = worksheets[3]

    for column in reversed(range(1, worksheet.max_column + 1)):
        if worksheet.cell(1, column).value == "Margin":
            worksheet.delete_cols(column)

    WORKBOOK.save(INPUT_FILE)
    print(f"{GREEN}Margin removed!{RESET}\n")
    print_sheets(3)

def main():
    """Interactive CLI with menu"""
    menu_options = {
        1: print_sheets,
        2: find_low_quantity,
        3: print_margin,
        4: write_margin,
        5: delete_margin,
        6: lambda: print_sheets(3)
    }
    print(f"Type ({RED}q{RESET}) to quit\n")
    while True:
        show_menu()
        try:
            user_input = input(f"Enter your choice (1-{len(menu_options)}): ")
            print()
            if user_input.lower() in ('q', 'quit', 'exit'):
                break
            user_input = int(user_input)
            if user_input in menu_options:
                menu_options[user_input]()
            print()
        except ValueError:
            continue
        except KeyboardInterrupt:
            print()
            break
        except EOFError:
            print()
            break

main()
